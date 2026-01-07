import csv

import openpyxl
from student import *
from openpyxl.styles import Font, Alignment, PatternFill

class StudentManager:
    def __init__(self):
        self.students = []

    # Thêm sinh viên
    def add_student(self, student):
        out = ""
        if self.find_student(student.id) == "Tìm thành công!\n":
            out += "Sinh viên đã tồn tại!\n"
            return out
        elif (student.gpa < 0.00) or (student.gpa > 4.00):
            out += "GPA không hợp lệ!\n"
            return out
        else :
            self.students.append(student)
            out += "Đã thêm " + str(student) + "thành công!\n"
            out += self.save_to_file() + "\n"
            return out

    # Hiển thị danh sách sinh viên, trả về chuỗi
    def show_students(self):
        out = ""
        if not self.students:
            out += "Danh sách sinh viên trống!\n"
            return out
        else:
            out += "Danh sách sinh viên:\n"
            for student in self.students:
                out += str(student) + "\n"
            return out

    # Tìm kiếm sinh viên theo ID hoặc tên, trả về chuỗi
    def find_student(self, student_id):
        out = ""
        list = [student for student in self.students 
                       if student_id == student.id or student_id == student.name]
        if list:
            out += "Tìm thành công!\n"
            for student in list:
                out += str(student) + "\n"
            return out
        else:
            out += "Sinh viên không tồn tại!\n"
            return out

    # Xóa sinh viên theo ID
    def delete_student(self, student_id):
        out = ""
        if not self.find_student(student_id):
            out += "Sinh viên không tồn tại!\n"
            return out
        else :
            self.students = [student for student in self.students if student.id != student_id]
            out += "Đã xóa thành công!\n"
            out += self.save_to_file() + "\n"
            return out

    # Lưu dữ liệu vào file CSV, trả về chuỗi
    def save_to_file(self, file_name="students.csv"):
        out = ""
        with open(file_name, mode="w", newline="", encoding="utf-8") as file: 
            writer = csv.writer(file)
            writer.writerow(["ID", "Name", "GPA", "Year"]) 
            for student in self.students: writer.writerow([student.id, student.name, student.gpa, student.year]) 
            out += "Đã lưu dữ liệu vào file " + file_name + "\n"
            return out

    # Load dữ liệu từ file CSV
    def load_from_file(self, file_name="students.csv"): 
        out = ""
        try: 
            with open(file_name, mode="r", encoding="utf-8") as file: 
                reader = csv.DictReader(file) 
                self.students = [Student(row["ID"], row["Name"], row["GPA"], row["Year"]) for row in reader] 
                out += "Đã tải dữ liệu từ file " + file_name + "\n"
                return out
        except FileNotFoundError: 
            out += "File không tồn tại!\n"
            return out
        
    # Thay đổi thông tin sinh viên
    def edit_infomation(self, student_id):
        out = ""
        if student_id == "0":
            return
        if not self.find_student(student_id):
            return
        else :
            name = input("Nhập tên mới: ")
            gpa = input("Nhập GPA mới: ")
            year = input("Nhập năm học mới: ")
            self.students = [Student(student_id, name, gpa, year) if s.id == student_id else s for s in self.students]
            out += "Đã sửa thành công!\n"
            out += self.save_to_file() + "\n"
            return out

    # Lưu ngay dữ liệu ra file Excel
    def export_to_excel(self, file_name="students.xlsx"):
        out = ""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Students"
        headers = ["ID", "Name", "GPA", "Year"]
        sheet.append(headers)
        for col in range(1, len(headers)+1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for student in self.students:
            sheet.append([student.id, student.name, student.gpa, student.year])
        wb.save(file_name)
        out += "Đã xuất dữ liệu ra file " + file_name + "\n"
        return out

    # Load -> Sắp xếp -> Lưu
    def sort_students_by_gpa(self):
        out = ""
        self.load_from_file()
        self.students.sort(key=lambda student: float(student.gpa), reverse=True)
        out += "Đã sắp xếp sinh viên theo GPA giảm dần.\n"
        out += self.save_to_file() + "\n"
        return out

    # Load -> Sắp xếp -> Lưu
    def sort_students_by_id(self):
        out = ""
        self.load_from_file()
        self.students.sort(key=lambda student: student.id)
        out += "Đã sắp xếp sinh viên theo mã số tăng dần.\n"
        out += self.save_to_file() + "\n"
        return out

    # Nhập dữ liệu từ file Excel
    def import_from_excel(self, file_name="input.xlsx"):
        out = ""
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            count_new = 0
            count_updated = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sid, name, gpa, year = row
                student = Student(sid, name, gpa, year)
                existing_student = self.find_student(sid)
                if existing_student:
                    self.students = [student if s.id == sid else s for s in self.students]
                    count_updated += 1
                else:
                    self.students.append(student)
                    count_new += 1
            out += f"Đã thêm {count_new} sinh viên mới và cập nhật {count_updated} sinh viên từ file Excel.\n"
            out += self.save_to_file() + "\n"
            return out
        except FileNotFoundError:
            out += "File Excel không tồn tại!\n"
        return out

    # Phân loại sinh viên theo GPA
    def classify_students_by_gpa(self, gpa):
        out = f"Sinh viên có GPA >= {gpa}:\n"
        classified_students = [student for student in self.students if float(student.gpa) >= gpa]
        if classified_students:
            for student in classified_students:
                out += str(student) + "\n"
        else:
            out += "Không có sinh viên nào đạt yêu cầu.\n"
        return out

    # Thống kê điểm
    def statistical_analysis(self):
        out = ""
        self.load_from_file()
        if not self.students:
            out += "Danh sách sinh viên trống!\n"
            return out
        gpas = [float(student.gpa) for student in self.students]
        average_gpa = sum(gpas) / len(gpas)
        max_gpa = max(gpas)
        min_gpa = min(gpas)
        stats = {
            "Trung bình": average_gpa,
            "Max": max_gpa,
            "Min": min_gpa,
            "Phương sai": sum((x - average_gpa) ** 2 for x in gpas) / len(gpas),
            "Độ lệch chuẩn": (sum((x - average_gpa) ** 2 for x in gpas) / len(gpas)) ** 0.5
        }
        out = "Thống kê điểm GPA:\n"
        for key, value in stats.items():
            out += f"{key}: {value:.2f}\n"
        return out