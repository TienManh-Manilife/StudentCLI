import csv
from turtle import st

import openpyxl
from student import *
from openpyxl.styles import Font, Alignment, PatternFill

class StudentManager:
    def __init__(self):
        self.students = []

    # Thêm sinh viên
    def add_student(self, student):
        if self.find_student(student.id):
            print("Sinh viên đã tồn tại!")
            return
        elif (student.gpa < 0.00) or (student.gpa > 4.00):
            print("GPA không hợp lệ! Phải trong khoảng 0.0 đến 4.0")
            return 
        else :
            self.students.append(student)
            print("Thêm sinh viên thành công!")
            self.save_to_file()
            return

    # Hiển thị danh sách sinh viên
    def show_students(self):
        if not self.students:
            print("Danh sách trống!")
        else:
            print("Danh sách:")
            for student in self.students:
                print(student)

    # Tìm kiếm sinh viên theo ID hoặc tên
    def find_student(self, student_id):
        list = [student for student in self.students 
                       if student_id == student.id or student_id == student.name]
        if list:
            for student in list:
                print("Tìm thành công!")
                print(student)
                return True
        else:
            print("Sinh viên không tồn tại!")
            return False

    # Xóa sinh viên theo ID
    def delete_student(self, student_id): 
        if not self.find_student(student_id):
            print("Sinh viên không tồn tại nên không thể xóa!")
            return
        else :
            self.students = [student for student in self.students if student.id != student_id]
            print("Đã xóa thành công!")
            self.save_to_file()

    # Lưu dữ liệu vào file CSV
    def save_to_file(self, file_name="students.csv"): 
        with open(file_name, mode="w", newline="", encoding="utf-8") as file: 
            writer = csv.writer(file)
            writer.writerow(["ID", "Name", "GPA", "Year"]) 
            for student in self.students: writer.writerow([student.id, student.name, student.gpa, student.year]) 
            print("Đã lưu dữ liệu vào file", file_name)

    # Load dữ liệu từ file CSV
    def load_from_file(self, file_name="students.csv"): 
        try: 
            with open(file_name, mode="r", encoding="utf-8") as file: 
                reader = csv.DictReader(file) 
                self.students = [Student(row["ID"], row["Name"], row["GPA"], row["Year"]) for row in reader] 
                print("Đã tải dữ liệu từ file", file_name) 
        except FileNotFoundError: print("Chưa có file dữ liệu, bắt đầu mới.")

    # Thay đổi thông tin sinh viên
    def edit_infomation(self, student_id):
        if student_id == "0":
            return
        if not self.find_student(student_id):
            return
        else :
            print("Nhập thông tin mới:")
            name = input("Nhập tên mới: ")
            gpa = input("Nhập GPA mới: ")
            year = input("Nhập năm học mới: ")
            self.students = [Student(student_id, name, gpa, year) if s.id == student_id else s for s in self.students]
            print("Đã sửa thành công!")
            self.save_to_file()

    # Lưu ngay dữ liệu ra file Excel
    def export_to_excel(self, file_name="students.xlsx"):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Students"
        headers = ["ID", "Name", "GPA", "Year"]
        sheet.append(headers)
        for col in range(1, len(headers)+1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for student in self.students:
            sheet.append([student.id, student.name, student.gpa, student.year])
        wb.save(file_name)
        print("Đã xuất dữ liệu ra file Excel:", file_name)

    # Load -> Sắp xếp -> Lưu
    def sort_students_by_gpa(self):
        self.load_from_file()
        self.students.sort(key=lambda student: float(student.gpa), reverse=True)
        print("Đã sắp xếp sinh viên theo GPA giảm dần.")
        self.show_students()
        self.save_to_file()

    # Load -> Sắp xếp -> Lưu
    def sort_students_by_id(self):
        self.load_from_file()
        self.students.sort(key=lambda student: student.id)
        print("Đã sắp xếp sinh viên theo mã số tăng dần.")
        self.show_students()
        self.save_to_file()

    # Nhập dữ liệu từ file Excel
    def import_from_excel(self, file_name="input.xlsx"):
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            count_new = 0
            count_updated = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sid, name, gpa, year = row
                student = Student(sid, name, gpa, year)
                existing_student = self.find_student(sid)
                if existing_student:
                    self.students = [student if s.id == sid else s for s in self.students]
                    count_updated += 1
                else:
                    self.students.append(student)
                    count_new += 1
            print(f"Đã thêm {count_new} sinh viên mới và cập nhật {count_updated} sinh viên từ file Excel.")
            self.save_to_file()
        except FileNotFoundError:
            print("File Excel không tồn tại!")